from openpyxl import Workbook

favorite_foods = ['Pizza', 'Chocolate Cake', 'Broccoli']
favorite_colors = ['Blue', 'Purple', 'Green', 'Orange']

# Make a new workbook
workbook = Workbook()

# Get the active sheet. Workbooks are created with one sheet
worksheet = workbook.active

worksheet.title = 'Favorite Things'

# Write a column title to cell row = 1, column = 1
# row and column counts are numbered from 1
worksheet.cell(1, 1, 'Favorite foods')

for index, food in enumerate(favorite_foods):
    worksheet.cell(index + 2, 1, food)


# Write favorite colors to the second column

# Write a column title to cell row = 1, column = 2
worksheet.cell(1, 2, 'Favorite colors')

for index, color in enumerate(favorite_colors):
    worksheet.cell(index + 2, 2, color)


# Save workbook
workbook.save('favorites.xlsx')





